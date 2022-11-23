from train import train_model
from utils.process import extract_mfcc,split_data
from settings import *
import argparse
import numpy as np
import os
from tensorflow import keras
from openpyxl import load_workbook
from model.Transformer import TransformerBlock
import tensorflow as tf
gpus = tf.config.experimental.list_physical_devices('GPU')
if gpus:
  try:
    tf.config.experimental.set_virtual_device_configuration(
        gpus[0],[tf.config.experimental.VirtualDeviceConfiguration(memory_limit=5120)])
  except RuntimeError as e:
    print(e)


parser = argparse.ArgumentParser()
parser.add_argument("--process",help="处理excel文件")
parser.add_argument("--train",help='训练模型')
parser.add_argument("--valid",help='预测模型')
parser.add_argument("--predict",help="模型预测")
parser.add_argument("--model_list",help="输入需要训练的模型,格式为 --model_list=CNN,FNN,LSTM,BiLSTM")

args = parser.parse_args()

if args.process:
    if args.process == "train":
        print('从文件提取训练集mfcc特征...')
        trainset,trainlabel = extract_mfcc('dataset/train.xlsx',PROCESSWITHNOISE)
        print(f'生成训练集个数: {len(trainset)}')
        np.save(os.path.join(NPYSAVEPATH,'train_set.npy'),trainset)
        np.save(os.path.join(NPYSAVEPATH,'train_label.npy'),trainlabel)
    if args.process == "valid":
        print('从文件提取训练集mfcc特征...')
        trainset,trainlabel = extract_mfcc('dataset/train.xlsx',False)
        print(f'生成训练集个数: {len(trainset)}')
        np.save(os.path.join(NPYSAVEPATH,'valid_set.npy'),trainset)
        np.save(os.path.join(NPYSAVEPATH,'valid_label.npy'),trainlabel)
    if args.process == "predict":   
        print('从文件提取测试集mfcc特征...')
        testset,testlabel = extract_mfcc('dataset/test.xlsx',PROCESSWITHNOISE)
        print(f'测试集个数: {len(testset)}')
        np.save(os.path.join(NPYSAVEPATH,'test_set.npy'),testset)
    
if args.train:
    if args.model_list:
        train_set = np.load(os.path.join(NPYSAVEPATH,'train_set.npy'))
        train_label = np.load(os.path.join(NPYSAVEPATH,'train_label.npy'))
        data = split_data(train_set, train_label, test_size=0.1, valid_size=0.1)
        model_list = args.model_list.split(',')
        trained_model_list = train_model(model_list,data,BATCHSIZE,PATIENCE,EPOCH)
        for item in trained_model_list:
            trained_model_list[item].save(os.path.join(MODELSAVEPATH,f"{item}.h5"))
            trained_model_list[item].summary()
    else:
        print('请输入model_list参数！')

if args.valid:
    valid_set = np.load(os.path.join(NPYSAVEPATH,'valid_set.npy'))
    valid_label = np.load(os.path.join(NPYSAVEPATH,'valid_label.npy'))
    trained_models = args.model_list.split(',')
    for item in trained_models:
        
        # print(item)
        if 'Transformer' in item:
            model = keras.models.load_model(os.path.join(MODELSAVEPATH,item+'.h5'),custom_objects={"TransformerBlock": TransformerBlock })
        else:
            model = keras.models.load_model(os.path.join(MODELSAVEPATH,item+'.h5'))
        temp = model.evaluate(np.array(valid_set),np.array(valid_label),verbose=1)
        print(f'{item}模型验证精度为{temp[1]}')
            



      
if args.predict:
    test_set = np.load(os.path.join(NPYSAVEPATH,'test_set.npy'))
    trained_models = args.model_list.split(',')
    predict_dict = {}
    sum_predict_arr = np.zeros((200,1))
    num = 0
    for item in trained_models:
        if 'Transformer' in item:
            model = keras.models.load_model(os.path.join(MODELSAVEPATH,item+'.h5'),custom_objects={"TransformerBlock": TransformerBlock })
        else:
            model = keras.models.load_model(os.path.join(MODELSAVEPATH,item+'.h5'))
        temp = model.predict(np.array(test_set))
        predict_dict[item] = np.around(temp)
        sum_predict_arr += np.around(temp)
        num += 1
    mean_predict_arr =np.around(sum_predict_arr/num)

    for item in trained_models:
        print(f'{item} Difference...')
        print(np.where(mean_predict_arr != predict_dict[item])[0])
   
    
    wb = load_workbook('dataset/test.xlsx')
    sheet = wb.active
    for i in range(len(mean_predict_arr)):
        sheet.cell(row=i+2, column=2).value = int(mean_predict_arr[i])
    wb.save('test_main.xlsx')
    